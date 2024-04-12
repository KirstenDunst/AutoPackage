# 合并国际化之后的excel文件，生成一个合并之后的新文件


import os
from openpyxl import Workbook, load_workbook


if __name__ == '__main__':
    base_excel_path = input('基础翻译文件excel地址:')
    base_excel_path = base_excel_path.removeprefix("'")
    base_excel_path = base_excel_path.removesuffix("'")
    add_excel_path = input('待合并的翻译excel文件地址:')
    add_excel_path = add_excel_path.removeprefix("'")
    add_excel_path = add_excel_path.removesuffix("'")
    excel_merge_param = input('待合并文件的列:基础文件的列,多个之间用英文逗号隔开,下标从0开始(eg: 1:1,2:3):')
    
    param_arr = excel_merge_param.split(',')
    merge_map:dict[int:int] = {}
    for item in param_arr:
        temp_arr = item.split(':')
        merge_map[temp_arr[0]] = temp_arr[-1]
    
    # 打开文件
    workbook = load_workbook(add_excel_path)
    # 获取所有工作表名称
    sheet_names = workbook.sheetnames
    sheet_index = 0
    if len(sheet_names) > 1:
        print(str(sheet_names))
        sheet_index = input('选择工作表的下标:')
    ws = workbook[sheet_names[int(sheet_index)]]
    
    # 中文(key): {excel哪个位置：将要填入取代的值}
    add_local_map:dict[str:dict[int:str]] = {}
    # 循环每一行
    for i in range(ws.max_row):
        value:dict[int:str] = {}
        key = ''
        for j in range(ws.max_column):
            cell = ws.cell(row=i+1, column=j+1)
            # print(cell.value)
            if (j == 0):
                # 去除excel表中文key字符串两端可能存在的空格
                key = cell.value.strip()
            else: 
                merge_map_key = str(j)
                if merge_map_key in merge_map:
                    value[int(merge_map[merge_map_key])] = cell.value
        if len(value) > 0:
            add_local_map[key] = value
            
    baseBook = load_workbook(base_excel_path)
    # 这里暂定只有一个 WS工作sheet，如果有需要再添加input询问
    baseWS = baseBook[baseBook.sheetnames[0]]
    outwb = Workbook()
    outws = outwb.worksheets[0]
    # 循环每一行
    for i in range(baseWS.max_row):
        line_content = []
        key = ''
        for j in range(baseWS.max_column):
            cell = baseWS.cell(row=i+1, column=j+1)
            if (j == 0):
                # 去除excel表中文key字符串两端可能存在的空格
                key = cell.value.strip()
                line_content.append(key)
            else: 
                if len(key) >0 and key in add_local_map and j in add_local_map[key]:
                    # 新增的内容覆盖之前的调整（红色提醒）
                    line_content.append(add_local_map[key][j])
                    add_local_map[key].pop(j)
                else:
                    line_content.append(cell.value)          
        if len(key) > 0 and key in add_local_map and len(add_local_map[key]) > 0:
            # 新增列
            last_line_map:dict[int, str] = add_local_map[key]
            max_column = baseWS.max_column
            index = baseWS.max_column
            for k, v in last_line_map.items():
                max_column = max(max_column,k)
            
            while index <= max_column:
                if index in last_line_map:
                    # 新增的列内容（橙色提醒）
                    line_content.append(last_line_map[index])
                else:
                    line_content.append('')
                index+=1     
        
        if len(line_content) > 0:
            outws.append(line_content)
        
        if key in add_local_map:
            add_local_map.pop(key)
            
    if len(add_local_map) > 0:
        for k,value in add_local_map.items():
            line_content = [k] 
            add_max_column = 0
            for key, v in value.items():
                add_max_column = max(add_max_column,key)
            
            add_index = 1
            while add_index <= add_max_column:
                if add_index in value:
                    # 新增的行内容（橙色提醒）
                    line_content.append(value[add_index])
                else:
                    # 新增的行内容（橙色提醒）
                    line_content.append('')
                add_index+=1     
            
            if len(line_content) > 0:
                outws.append(line_content)
    
    
    save_path = base_excel_path.replace('.xlsx','_merge.xlsx')
    if os.path.exists(save_path):
        os.remove(save_path)
    outwb.save(save_path)
        