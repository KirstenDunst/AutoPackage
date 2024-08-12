# excel生成showdoc文档excel格式，和showdoc中excel格式生成excel

import os

from openpyxl import Workbook, load_workbook


class ShowDocExcelTool:
    
    """
    将excel转化成showdoc的excel格式
    """
    @staticmethod
    def transExcelToDoc(excel_file_path:str,doc_generate_path:str):
        workbook = load_workbook(excel_file_path)
        sheet_index = 0
        if len(workbook.sheetnames) > 1:
            print(str(workbook.sheetnames))
            sheet_index = input('选择工作表的下标:')
        ws = workbook[workbook.sheetnames[int(sheet_index)]]
        
        showdoc_content = ''
        for i in range(ws.max_row):
            space_content = '|'
            cell_content = '|'
            for j in range(ws.max_column):
                cell = ws.cell(row=i+1, column=j+1)
                temp_value = str(cell.value)
                temp_value.replace('\n','\\n')
                space_content += (temp_value + '|')
                #居中对齐
                cell_content+=':-:|'
            
            showdoc_content+= space_content
            showdoc_content+='\n'
            if i == 0:
                showdoc_content+= cell_content
                showdoc_content+='\n'
    
        file = open(doc_generate_path,'w')
        file.write(showdoc_content) #写入内容信息
        file.close()
        
    """
    将showdoc的excel格式转化成excel
    """
    @staticmethod
    def transDocToExcel(doc_file_path:str,excel_generate_path:str):
        
        outwb = Workbook()
        outws = outwb.worksheets[0] 
    #     outws.append(excel_title)
        
    # for k, v in trans_dict['zh_CN'].items():
    #     a_list = [v]
    #     for language_code in language_code_arr[1:]:
    #         a_list.append(trans_dict[language_code][k])
    #     outws.append(a_list)       
            
    # if os.path.exists(out_excel_file):
    #     os.remove(out_excel_file)
    # outwb.save(out_excel_file)


if __name__ == '__main__':
    
    # excel 2 showdoc
    excel_path = input('需要转为showdoc文件的excel地址:')
    excel_path = excel_path.removeprefix("'")
    excel_path = excel_path.removesuffix("'")
    
    save_showdoc_path = os.path.dirname(excel_path)+'/showdoc.text'
    if os.path.exists(save_showdoc_path):
        ensure = input('当前excel文件所在目录下存在showdoc.text文件是否覆盖?[O:覆盖,Q:中断]')
        if ensure.upper() != 'O':
            exit()
        os.remove(save_showdoc_path)
    ShowDocExcelTool.transExcelToDoc(excel_path,save_showdoc_path)
    
    
    # # showdoc  2 excel
    # generate_excel_parent_path = input('生成excel保存地址的父级文件夹:')
    # generate_excel_parent_path = generate_excel_parent_path.removeprefix("'")
    # generate_excel_parent_path = generate_excel_parent_path.removesuffix("'")
    
    # showdoc_path = generate_excel_parent_path+'/showdoc.text'
    # if os.path.exists(showdoc_path):
    #     os.remove(showdoc_path)
    # file = open(showdoc_path,'w')
    # file.write("") #写入内容信息
    # file.close()
        
    # is_continue = input('已创建showdoc.text,填写完成后输入C以继续:')
    # if not is_continue.upper() == 'C':
    #     os.remove(showdoc_path)
    #     exit('您已中断执行')
    # ShowDocExcelTool.transDocToExcel(showdoc_path,generate_excel_parent_path+'/generate.xlsx')
    
    
    
    