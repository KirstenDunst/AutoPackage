# 两个国际化的excel的对比处理，生成项目缺少翻译的excel表

import math
import os
import string
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill


# 如上图,一共10列,从0到9
# 产生为10的excel对应的列
def cycle_letter(arr,level):
    list1 = string.ascii_uppercase
    tempArr = []
    letterArr = [i for i in list1]
    arrNum = len(arr)
    if(level==0 or arrNum==0):
        return letterArr
    for index in range(arrNum):
        for letter in letterArr:
            tempArr.append(arr[index]+letter)
    return tempArr
 
def reduce_excel_col_name(num):
    tempVal = 1
    level = 1
    while(tempVal):
        tempVal = num/(math.pow(26, level))
        if(tempVal>1):
            level += 1
        else:
            break
    excelArr = []
    tempArr = []
    for index in range(level):
        tempArr = cycle_letter(tempArr,index)
        for numIndex in range(len(tempArr)): 
            if(len(excelArr)<num):
                excelArr.append(tempArr[numIndex])
            else:
                return excelArr
    return excelArr

if  __name__ == '__main__':
    has_trans_excel_path = input('请输入已经国际化的excel路径:')
    has_trans_excel_path = has_trans_excel_path.removeprefix("'")
    has_trans_excel_path = has_trans_excel_path.removesuffix("'")
    
    project_excel_path = input('请输入项目导出的国际化excel路径:')
    project_excel_path = project_excel_path.removeprefix("'")
    project_excel_path = project_excel_path.removesuffix("'")
    
    pair_param = input('项目纵列与已翻的列对应关系(0:中文都在第一行,不用写对应关系):[1:1,2:3,3:2]:')
    param_arr = pair_param.split(',')
    pair_map:dict[int:int] = {}
    total_param_max_column = 0
    for item in param_arr:
        if item == '0:0':
            continue
        temp_arr = item.split(':')
        project_column = int(temp_arr[0])
        total_column = int(temp_arr[-1])
        total_param_max_column = max(total_param_max_column,total_column)
        pair_map[project_column] = total_column
    
    
    # 打开文件
    trans_workbook = load_workbook(has_trans_excel_path)
    trans_ws = trans_workbook[trans_workbook.sheetnames[0]]
    
    if total_param_max_column+2 > trans_ws.max_column :
        exit('列匹配,已翻译总表列数读取超出:总表excel最大列数'+str(trans_ws.max_column))
    
    # 中文(key): {excel哪个位置：当前的值}
    trans_map:dict[str:dict[int:str]] = {}
    # 循环每一行
    for i in range(trans_ws.max_row):
        value:dict[int:str] = {}
        key = ''
        for j in range(trans_ws.max_column):
            cell = trans_ws.cell(row=i+1, column=j+1)
            if (j == 0):
                key = cell.value
            else: 
                # None 处理设置默认值''
                value[j] = cell.value if cell.value else ''
        if i > 0 :
            trans_map[key]= value
    
    outwb = Workbook()
    outws = outwb.worksheets[0]
    
    project_workbook = load_workbook(project_excel_path)
    project_ws = project_workbook[project_workbook.sheetnames[0]]
    
    # 标题
    titles = [project_ws.cell(row=1, column=1).value]
    for k in range(project_ws.max_column):
        cell = project_ws.cell(row=1, column=k+1)
        if k in pair_map :
            titles.append(cell.value)
    outws.append(titles)
    
    column_index_map = {1:'A',2:'B',3:'C',4:'D',5:'E',6:'F',7:'G',8:'H',9:'I',10:'J',11:'K',12:'L',13:'M',14:'N',15:'O',16:'P',17:'Q',18:'R',19:'S',20:'T',21:'U',22:'V',23:'W',24:'X',25:'Y',26:'Z',27:'AA',28:'AB',29:'AC'}

    color_indexs = []
    color_row = 2
    # 循环每一行(从标题下面那行开始循环)
    for i in range(1, project_ws.max_row):
        arr:list[str] = []
        color_column_arr = []
        key = ''
        for j in range(project_ws.max_column):
            cell = project_ws.cell(row=i+1, column=j+1)
            if (j == 0):
                key = cell.value
            else: 
                if j in pair_map :
                    # 未翻译的用空字符占位存放在excel里面
                    cell_value = ''
                    trans_row = trans_map[key] if key in trans_map else {}
                    if pair_map[j] in trans_row:
                        cell_value = trans_row[pair_map[j]]
                        
                    if cell_value != '':
                        color_column_arr.append(column_index_map[j+1])
                        
                    arr.append(cell_value)
            
        no_empty_arr = [num for num in arr if num != '']   
        if len (arr) > 0 and len(no_empty_arr) != len(arr): 
            for cell in color_column_arr:
                color_indexs.append(cell+str(color_row))
            color_row += 1
            # 此行中文 有未翻译的语种
            arr.insert(0,key) 
            outws.append(arr)
    
    save_path = project_excel_path.replace('.xlsx','_need_trans.xlsx')
    if os.path.exists(save_path):
        os.remove(save_path)
    outwb.save(save_path)
    
    if len(color_indexs) > 0 :
        # 颜色填充
        fill = PatternFill('solid',fgColor='FFFF00')
        wb = load_workbook(save_path)
        work = wb[wb.sheetnames[0]]
        print(len(color_indexs))
        for i in color_indexs:
            work[i].fill = fill
        wb.close()
        wb.save(save_path)
    
    
    