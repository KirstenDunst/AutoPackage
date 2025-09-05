import os
import json
from openpyxl import load_workbook
from openpyxl import Workbook

from common_tool import CommonTool


class ExcelTool:
    # 转换
    @staticmethod
    def trDict(sourceDict: dict, cloumnIndex: int, dataMap: dict):
        newDict = {}
        for k, v in sourceDict.items():
            if str(v.__class__).__contains__("dict"):
                dealDict = ExcelTool.trDict(v, cloumnIndex, dataMap)
                newDict.update({k: dealDict})
            else:
                if dataMap.keys().__contains__(v):
                    values: list[str] = dataMap[v]
                    if len(values) >= cloumnIndex:
                        newDict[k] = values[cloumnIndex - 1]
                    else:
                        newDict[k] = ""
                else:
                    newDict[k] = ""
        return newDict

    # 对比json，将主json对应value为''的辅json结构保存下来返回
    @staticmethod
    def compareDict(mainDict: dict, subDict: dict):
        newDict = {}
        for k, v in mainDict.items():
            if str(v.__class__).__contains__("dict"):
                dealDict = ExcelTool.compareDict(v, subDict[k])
                newDict.update({k: dealDict})
            else:
                if v == "":
                    newDict[k] = subDict[k]
                    print(subDict[k])
        return newDict

    # 合并json, 新产生的空数据不合并, 否则同样的key，使用新json的value做value
    @staticmethod
    def mergeJson(oldDict: dict, newDict: dict):
        endDict = {}
        for k, v in newDict.items():
            if str(v.__class__).__contains__("dict"):
                if str(oldDict[k].__class__).__contains__("dict"):
                    dealDict = ExcelTool.mergeJson(oldDict[k], newDict[k])
                    endDict[k] = dealDict
                else:
                    endDict[k] = newDict[k]
            else:
                if oldDict.keys().__contains__(k):
                    # 新值包含，替换使用新值，如果新值没有则用旧值
                    if newDict[k] != "" and newDict[k] != None:
                        endDict[k] = newDict[k]
                    else:
                        endDict[k] = oldDict[k]
                else:
                    endDict[k] = v
            if oldDict.keys().__contains__(k):
                oldDict.pop(k)
        endDict.update(oldDict)
        return endDict

    # 判断一个字典是否有有效内容
    @staticmethod
    def isNoEmpty(tempDict: dict):
        hasData = len(tempDict) > 0
        if hasData:
            for k, v in tempDict.items():
                if str(v.__class__).__contains__("dict"):
                    hasData = ExcelTool.isNoEmpty(v)
                else:
                    hasData = True
        return hasData


if __name__ == "__main__":
    excel_file_path = input("翻译之后的excel文件地址:")
    excel_file_path = excel_file_path.removeprefix("'")
    excel_file_path = excel_file_path.removesuffix("'")
    chinese_locale_file = input("项目中国际化中文json文件地址:")
    chinese_locale_file = chinese_locale_file.removeprefix("'")
    chinese_locale_file = chinese_locale_file.removesuffix("'")
    locales_names = input(
        "将需要转化对应的列与对应的json命名(多个用,隔开)下标从0开始:(eg: 1:en_US,0:zh_CN):"
    )
    # 1:en_US,2:de_DE,3:es_ES,4:fr_FR,5:pt_PT,6:it_IT,7:pl_PL,8:ja_JP,9:nl_NL
    miss_file_parent_path = input("缺失翻译的文件整理存放父级文件夹目录:")
    miss_file_parent_path = miss_file_parent_path.removeprefix("'")
    miss_file_parent_path = miss_file_parent_path.removesuffix("'")

    # 存放生成
    root_path = os.path.abspath(os.path.dirname(__file__) + os.path.sep + "..")
    source_dir = root_path + "/assets/locales/"
    # 打开文件
    workbook = load_workbook(excel_file_path)
    # 获取所有工作表名称
    sheet_names = workbook.sheetnames
    sheet_index = 0
    if len(sheet_names) > 1:
        print(">>>>>>>>" + str(sheet_names))
        sheet_index = input("选择工作表的下标(从0开始):")
    ws = workbook[sheet_names[int(sheet_index)]]

    chinese_to_local_map: dict[str : list[str]] = {}
    # 循环每一行
    for i in range(ws.max_row):
        value: list[str] = []
        key = ""
        for j in range(ws.max_column):
            cell = ws.cell(row=i + 1, column=j + 1)
            # print(cell.value)
            if j == 0:
                key = cell.value
            else:
                value.append(cell.value)
        chinese_to_local_map[key] = value

    # 读取中文国际化json
    with open(chinese_locale_file, "r") as f:
        content = f.read()
    # 文件中的json
    json_content: dict = json.loads(content)

    save_file_arr = locales_names.split(",")
    file_map: dict = {}
    comapre_main_dict: dict = {}
    for item in save_file_arr:
        temp_arr = item.split(":")
        index = int(temp_arr[0])
        file_name = temp_arr[-1]
        file_content = ExcelTool.trDict(json_content, index, chinese_to_local_map)
        comapre_main_dict = file_content
        file_map[file_name] = file_content

    # 保存没有翻译的json待定（等下一次提供给专业翻译人员）, 这里目前只取第一个做对比即可
    no_trans_chinese = ExcelTool.compareDict(comapre_main_dict, json_content)

    hasData = ExcelTool.isNoEmpty(no_trans_chinese)
    if hasData:
        # title_map = {'en_US':'英文','de_DE':'德语','es_ES':'西班牙语','fr_FR':'法语','pt_PT':'葡萄牙语','it_IT':'意大利语','pl_PL':'波兰语','zh_CN':'简体中文','nl_NL':'荷兰语','ja_JP':'日语'}
        outwb = Workbook()
        outws = outwb.worksheets[0]
        compare_file_map: dict = {}
        titles = ["zh_CN"]
        for k, v in file_map.items():
            titles.append(k)
            # 这里是获取本地json里面的直接字典
            file_path = source_dir + k + ".json"
            temp_json_content = {}
            if os.path.exists(file_path):
                with open(file_path, "r") as f:
                    json_content = f.read()
                temp_json_content = json.loads(json_content)
            compare_file_map[k] = CommonTool.tranformDict(temp_json_content, "")
        outws.append(titles)

        direct_chinese_dict = CommonTool.tranformDict(no_trans_chinese, "")
        for k, v in direct_chinese_dict.items():
            a_list = [v]
            for name in titles[1:]:
                dic = compare_file_map[name]
                if k in dic:
                    a_list.append(dic[k])
            outws.append(a_list)
        save_path = miss_file_parent_path + "/miss_transform.xlsx"
        if os.path.exists(save_path):
            os.remove(save_path)
        outwb.save(save_path)

    # 文件写入
    for k, value in file_map.items():
        file_path = source_dir + k + ".json"
        if os.path.exists(file_path):
            skip = input(
                "检测到文件已存在:"
                + file_path
                + "跳过「S」?还是覆盖「O」?,还是合并「M」,或其他字符退出执行:"
            )
            skip_upper = skip.upper()
            if skip_upper == "S":
                continue
            elif skip_upper == "O":
                json_data = json.dumps(value, ensure_ascii=False, indent=2)
                with open(file_path, "w") as f:
                    f.write(json_data)
            elif skip_upper == "M":
                # 将excel文件中的新对应的value有内容的内容，和原有的json合并之后再写入文件
                # 读取原数据，合并json
                with open(file_path, "r") as f:
                    json_content = f.read()
                temp_json_content: dict = json.loads(json_content)
                temp_json_content = ExcelTool.mergeJson(temp_json_content, value)
                # 写入文件
                merge_json_data = json.dumps(
                    temp_json_content, ensure_ascii=False, indent=2
                )
                with open(file_path, "w") as f:
                    f.write(merge_json_data)
            else:
                exit("手动中断脚本")
        else:
            json_data = json.dumps(value, ensure_ascii=False, indent=2)
            with open(file_path, "w") as f:
                f.write(json_data)
