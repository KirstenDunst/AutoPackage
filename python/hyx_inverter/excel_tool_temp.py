# 临时处理国际化操作，无通用意义，直接改

import json
import os
from openpyxl import Workbook, load_workbook


def check_fr():
    # 将法语从excel中摘出来，带中文
    excel_file = input("excel文件路径:")
    excel_file = excel_file.removeprefix("'")
    excel_file = excel_file.removesuffix("'")
    # 确认过的列index
    ensure_index = 7

    # 打开文件
    workbook = load_workbook(excel_file, data_only=True)
    ws = workbook["APP"]
    outwb = Workbook()
    outws = outwb.worksheets[0]
    for i in range(ws.max_row):
        line_content = []
        line_content.append(ws.cell(row=i + 1, column=1).value)
        fr_content = ws.cell(row=i + 1, column=ensure_index).value
        if fr_content == None or len(fr_content) == 0:
            fr_content = ws.cell(row=i + 1, column=ensure_index - 1).value
        line_content.append(fr_content)

        outws.append(line_content)

    save_path = excel_file.replace(".xlsx", "_new.xlsx")
    if os.path.exists(save_path):
        os.remove(save_path)
    outwb.save(save_path)


def merge_release_json_fr_excel():
    release_json_path = input("线上运维平台json地址:")
    release_json_path = release_json_path.removeprefix("'")
    release_json_path = release_json_path.removesuffix("'")
    fr_excel_path = input("法语待合并excel地址:")
    fr_excel_path = fr_excel_path.removeprefix("'")
    fr_excel_path = fr_excel_path.removesuffix("'")

    ch_fr_map = {}
    workbook = load_workbook(fr_excel_path, data_only=True)
    ws = workbook[workbook.sheetnames[0]]
    for i in range(ws.max_row):
        chinese_value = ws.cell(row=i + 1, column=1).value
        language_value = ws.cell(
            row=i + 1, column=2
        ).value
        if chinese_value != None:
            if chinese_value in ch_fr_map:
                print(chinese_value)
            ch_fr_map[chinese_value] = language_value
    print(len(ch_fr_map))
    # print(ch_fr_map)
    with open(release_json_path, "r") as f:
        content = f.read()
    # 文件中的json
    json_content: dict = json.loads(content)
    f.close()

    for k, v in json_content.items():
        chinese = v["local"]["zh"]
        if chinese in ch_fr_map:
            # print(chinese)
            v["local"]["fr"] = ch_fr_map[chinese]

    save_path = release_json_path.replace(".json", "_new.json")
    if os.path.exists(save_path):
        os.remove(save_path)
    with open(save_path, "w") as f2:
        f2.write(json.dumps(json_content, ensure_ascii=False, indent=2))
    f2.close()


def local_locales_base_mergin_pingtai_remark():
    # 以本地json为主，合并平台json，如果本地json有平台json的key，那么将平台的对应key的remark拿过来,本地没有的平台key不会被加入
    # 旨在以本地为主，只是拿取平台的对应词组的remark
    local_json_path = input("本地项目生成的国际化json路径:")
    local_json_path = local_json_path.removeprefix("'")
    local_json_path = local_json_path.removesuffix("'")
    pingtai_json_path = input("平台全量导出的国际化json路径:")
    pingtai_json_path = pingtai_json_path.removeprefix("'")
    pingtai_json_path = pingtai_json_path.removesuffix("'")

    with open(pingtai_json_path, "r") as f2:
        content2 = f2.read()
    pingtai_json_content: dict = json.loads(content2)
    f2.close()

    pingtai_remark_map = {}
    for k, v in pingtai_json_content.items():
        pingtai_remark_map[k] = v["remark"] if "remark" in v else ""

    with open(local_json_path, "r") as f1:
        content = f1.read()
    local_json_content: dict = json.loads(content)
    f1.close()

    for k1, v1 in local_json_content.items():
        if k1 in pingtai_remark_map:
            v1["remark"] = pingtai_remark_map[k1]

    save_path = local_json_path.replace(".json", "_new.json")
    if os.path.exists(save_path):
        os.remove(save_path)
    with open(save_path, "w") as f3:
        f3.write(json.dumps(local_json_content, ensure_ascii=False, indent=2))
    f3.close()


def pingtai_locales_base_mergin_local_locales():
    # 以平台json为主，保留词汇的remark，对应key的locale国际化使用本地国际化替换，平台json没有的key不会被加入
    # 旨在平台为主，替换使用本地对应key的国际化
    local_json_path = input("本地项目生成的国际化json路径:")
    local_json_path = local_json_path.removeprefix("'")
    local_json_path = local_json_path.removesuffix("'")
    pingtai_json_path = input("平台全量导出的国际化json路径:")
    pingtai_json_path = pingtai_json_path.removeprefix("'")
    pingtai_json_path = pingtai_json_path.removesuffix("'")

    with open(local_json_path, "r") as f1:
        content = f1.read()
    local_json_content: dict = json.loads(content)
    f1.close()

    local_locales_map = {}
    for k, v in local_json_content.items():
        local_locales_map[k] = v["local"]

    with open(pingtai_json_path, "r") as f2:
        content2 = f2.read()
    pingtai_json_content: dict = json.loads(content2)
    f2.close()

    for k1, v1 in pingtai_json_content.items():
        if k1 in local_locales_map:
            v1["local"] = local_locales_map[k1]

    save_path = pingtai_json_path.replace(".json", "_new.json")
    if os.path.exists(save_path):
        os.remove(save_path)
    with open(save_path, "w") as f3:
        f3.write(json.dumps(pingtai_json_content, ensure_ascii=False, indent=2))
    f3.close()


if __name__ == "__main__":

    # # 法语摘出（注意需要手动调整soc和soh的大小写）
    # check_fr()

    # # 合并线上运维平台国际化json和法语国际化
    # merge_release_json_fr_excel()

    # # 国际化以本地为主合并平台remark
    # local_locales_base_mergin_pingtai_remark()

    # 平台为主，使用本地对应key的国际化
    pingtai_locales_base_mergin_local_locales()
